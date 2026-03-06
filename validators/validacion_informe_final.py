# ================================================================
# Archivo: validators/validacion_informe_final.py
# Descripción:
#   Genera el informe final consolidado con:
#     - Clasificados (nuevos)
#     - Previamente Procesado (Erroneos_Ultima_Carga_Titulares)
#     - Rechazo x MCredito (mayores 1200)
#     - RechazadosOtros (filtrado por RUT procesados)
#     - Resumen Detalle (formato institucional)
#
# Versión: v3.3.3 - 22/10/2025
# Autor: José López
# Institución: TRANSFORMACION DIGITAL
# ================================================================

import os
import glob
import warnings
import pandas as pd
import pyodbc
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

warnings.filterwarnings("ignore", category=UserWarning)

# ================================================================
# CONEXIÓN A BASE DE DATOS PRODUCCION
# ================================================================
CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=wpvmh-replica\\instancia1;"
    "DATABASE=PRODUCCION_migra;"
    "Authentication=ActiveDirectoryIntegrated;"
    "Encrypt=yes;"
    "TrustServerCertificate=yes;"
)

EXCEL_RECHAZADOS = (
    r"Z:\Negocio\Rebaja De Dividendo\DS02\Clasificacion DS02_DS116_DS120 & DS19\Clasificados & Rechazados DS02_DS116_DS19 & DS120.xlsx"
)

# ================================================================
# FUNCIONES DE APOYO
# ================================================================
def _formatear_libro(ruta_excel):
    wb = load_workbook(ruta_excel)
    borde = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for ws in wb.worksheets:
        if ws.max_row == 1 and ws.max_column == 1 and not ws["A1"].value:
            continue
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                c.border = borde
                c.alignment = Alignment(horizontal="left")
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 65)
        ws.freeze_panes = "A2"
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass
    wb.save(ruta_excel)


def _obtener_datos_sql(ruts):
    if not ruts:
        return pd.DataFrame(columns=["RUT", "DV", "Programa", "Región del Subsidio"])
    rut_str = ",".join([f"'{r}'" for r in ruts])

    q = f"""
    SELECT PER.PER_RUT AS RUT, PER.PER_DIG AS DV, PRG.PRG_NOM AS Programa,
           ISNULL(REG.REG_DES,'') AS [Región del Subsidio]
    FROM SUBSIDIO SUB
    INNER JOIN LINEA_DE_PROCESO LDP ON LDP.LIN_PRO_ID = SUB.LIN_PRO_ID
    INNER JOIN PROGRAMA PRG ON PRG.PRG_ID = LDP.PRG_ID
    INNER JOIN PERSONA PER ON SUB.IDO_ID = PER.IDO_ID
    LEFT JOIN REGION REG ON REG.REG_ID = SUB.REG_ID
    WHERE PER.PER_RUT IN ({rut_str})
    """
    conn = pyodbc.connect(CONN_STR)
    df = pd.read_sql(q, conn)
    conn.close()
    return df


def _buscar_archivo_erroneos(carpeta):
    patrones = [
        os.path.join(carpeta, "Erroneos_Ultima_Carga_Titulares*.xls"),
        os.path.join(carpeta, "Erroneos_Ultima_Carga_Titulares*.xlsx"),
    ]
    archivos = []
    for p in patrones:
        archivos.extend(glob.glob(p))
    if not archivos:
        return None
    archivos.sort(key=os.path.getmtime, reverse=True)
    return archivos[0]


def _leer_erroneos(path):
    if not path:
        return pd.DataFrame(columns=["RUT", "DV", "Motivo"])
    ext = os.path.splitext(path)[1].lower()
    try:
        df = pd.read_excel(path, engine="xlrd" if ext == ".xls" else "openpyxl")
    except Exception:
        df = pd.read_excel(path)
    df.columns = [c.strip().upper() for c in df.columns]
    rut_col = next((c for c in df.columns if "RUT" in c), None)
    dv_col = next((c for c in df.columns if c in ["DV", "DIGITO", "DÍGITO", "DIGITO VERIFICADOR"]), None)
    mot_col = next((c for c in df.columns if any(x in c for x in ["MOTIVO", "OBSERV", "ERROR", "DETALLE"])), None)

    out = pd.DataFrame()
    if rut_col:
        out["RUT"] = df[rut_col].astype(str).str.extract(r"(\d+)")[0]
    else:
        return pd.DataFrame(columns=["RUT", "DV", "Motivo"])
    out["DV"] = df[dv_col].astype(str).str.strip() if dv_col else ""
    out["Motivo"] = df[mot_col].astype(str) if mot_col else ""
    return out.drop_duplicates(subset=["RUT"])


def _leer_txt(path_txt):
    df = pd.read_csv(path_txt, sep=";", header=None, dtype=str, encoding="latin-1")
    df.fillna("", inplace=True)
    df.columns = range(df.shape[1])
    df["RUT"] = df[0].astype(str).str.strip()
    df["DV"] = df[1].astype(str).str.strip()
    df["Nombre"] = (df[2] + " " + df[3] + " " + df[4]).str.replace(r"\s+", " ", regex=True).str.strip()
    df["Monto UF"] = df[11].astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False).astype(float)

    def _porc(m):
        if m <= 500:
            return 20
        if m <= 900:
            return 15
        if m <= 1200:
            return 10
        return 0

    df["% Beneficio"] = df["Monto UF"].apply(_porc)
    return df[["RUT", "DV", "Nombre", "Monto UF", "% Beneficio"]]


def _generar_resumen(ruta_excel):
    wb = load_workbook(ruta_excel)
    names = wb.sheetnames

    df_clas = pd.read_excel(ruta_excel, sheet_name="Clasificados") if "Clasificados" in names else pd.DataFrame()
    df_prev = pd.read_excel(ruta_excel, sheet_name="Previamente Procesado") if "Previamente Procesado" in names else pd.DataFrame()
    df_rech = pd.read_excel(ruta_excel, sheet_name="Rechazo x MCredito") if "Rechazo x MCredito" in names else pd.DataFrame()
    df_otro = pd.read_excel(ruta_excel, sheet_name="RechazadosOtros") if "RechazadosOtros" in names else pd.DataFrame()

    ws = wb.create_sheet("Resumen Detalle")
    ws["A1"], ws["A2"] = "Buenos días XXXXX", "Envío detalle de clasificaciones procesadas"
    ws["A1"].font, ws["A2"].font = Font(bold=True, size=12), Font(size=11)

    start = 4
    ws[f"A{start}"], ws[f"B{start}"] = "Programa", "Deudores"
    programas = sorted(df_clas["Programa"].dropna().unique().tolist()) if not df_clas.empty else []
    row = start + 1
    for p in programas:
        cnt = df_clas[df_clas["Programa"] == p]["RUT"].nunique()
        ws[f"A{row}"], ws[f"B{row}"] = p, cnt
        row += 1
    ws[f"A{row}"], ws[f"B{row}"] = "Total", df_clas["RUT"].nunique() if not df_clas.empty else 0
    row += 3

    ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"], ws[f"E{row}"] = "Programa", "Deudores", "10%", "15%", "20%"
    row += 1
    resumen = []
    for p in programas:
        sub = df_clas[df_clas["Programa"] == p]
        d_total = len(sub)
        d10 = len(sub[sub["% Beneficio"] == 10])
        d15 = len(sub[sub["% Beneficio"] == 15])
        d20 = len(sub[sub["% Beneficio"] == 20])
        ws.append([p, d_total, d10, d15, d20])
        resumen.append((d_total, d10, d15, d20))

    tb = sum(r[0] for r in resumen)
    s10, s15, s20 = sum(r[1] for r in resumen), sum(r[2] for r in resumen), sum(r[3] for r in resumen)
    ws.append(["Sub - Total con Beneficio", tb, s10, s15, s20])
    ws.append(["Rechazo x MCredito", len(df_rech)])
    ws.append(["Previamente Procesado", len(df_prev)])
    ws.append(["RechazadosOtros", len(df_otro)])
    ws.append(["Total Detalle DS 116, DS19, DS 40 y DS 01", tb + len(df_rech) + len(df_prev) + len(df_otro)])

    wb.save(ruta_excel)
    _formatear_libro(ruta_excel)

# ================================================================
# FUNCIÓN PRINCIPAL
# ================================================================
def validar_informe_final(archivo_txt):
    try:
        carpeta = os.path.dirname(archivo_txt)
        base = os.path.basename(archivo_txt)
        path_men = archivo_txt.replace("_mayores_1200", "_menores_1200")
        path_may = archivo_txt.replace("_menores_1200", "_mayores_1200")
        nombre_salida = os.path.splitext(base)[0].replace("_menores_1200", "").replace("_mayores_1200", "")
        ruta_excel = os.path.join(carpeta, f"{nombre_salida}_InformeFinal.xlsx")

        # Crear hoja TEMP
        wb = Workbook()
        ws = wb.active
        ws.title = "TEMP"
        ws["A1"] = "Inicialización"
        wb.save(ruta_excel)

        # Clasificados y Previamente Procesado
        if os.path.exists(path_men):
            df_men = _leer_txt(path_men)
            df_men = df_men.merge(_obtener_datos_sql(df_men["RUT"].unique().tolist()), on="RUT", how="left")
            path_err = _buscar_archivo_erroneos(carpeta)
            df_err = _leer_erroneos(path_err)
            if not df_err.empty:
                mask = df_men["RUT"].isin(df_err["RUT"])
                df_prev = df_men[mask].merge(df_err, on="RUT", how="left")
                df_clas = df_men[~mask]
            else:
                df_prev = pd.DataFrame()
                df_clas = df_men
            with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a") as writer:
                if not df_clas.empty:
                    df_clas.to_excel(writer, index=False, sheet_name="Clasificados")
                if not df_prev.empty:
                    df_prev.to_excel(writer, index=False, sheet_name="Previamente Procesado")

        # Rechazo x MCredito
        if os.path.exists(path_may):
            df_may = _leer_txt(path_may)
            df_may["% Beneficio"] = 0
            df_may = df_may.merge(_obtener_datos_sql(df_may["RUT"].unique().tolist()), on="RUT", how="left")
            with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a") as writer:
                df_may.to_excel(writer, index=False, sheet_name="Rechazo x MCredito")

        # RechazadosOtros (filtrados por RUT procesados)
        rut_total = set()
        if os.path.exists(ruta_excel):
            xls = pd.ExcelFile(ruta_excel)
            for h in ["Clasificados", "Previamente Procesado", "Rechazo x MCredito"]:
                if h in xls.sheet_names:
                    rut_total |= set(pd.read_excel(ruta_excel, sheet_name=h)["RUT"].astype(str).tolist())

        if os.path.exists(EXCEL_RECHAZADOS) and rut_total:
            try:
                df_maestro = pd.read_excel(EXCEL_RECHAZADOS)
                df_maestro.columns = [c.strip() for c in df_maestro.columns]
                df_ro = df_maestro[
                    (df_maestro.get("Estado", "").astype(str).str.upper() == "RECHAZADO")
                    & (df_maestro.get("Rut", "").astype(str).isin(list(rut_total)))
                ][["Rut", "Dv", "Programa", "Región", "Motivo de Rechazo", "Monto Credito"]].copy()
                if not df_ro.empty:
                    df_ro.rename(columns={"Rut": "RUT", "Dv": "DV", "Región": "Región del Subsidio"}, inplace=True)
                    with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a") as writer:
                        df_ro.to_excel(writer, index=False, sheet_name="RechazadosOtros")
            except Exception as e:
                print(f"[AVISO] No se pudo leer el archivo institucional de rechazados: {e}")
                pass

        # Eliminar TEMP si hay otras hojas
        wb = load_workbook(ruta_excel)
        if "TEMP" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["TEMP"])
            wb.save(ruta_excel)

        _generar_resumen(ruta_excel)
        messagebox.showinfo("Informe generado", f"✅ Informe final creado correctamente:\n{ruta_excel}")
        return f"Informe generado: {ruta_excel}"

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar el informe:\n{e}")
        print(f"❌ Error en validar_informe_final: {e}")
        return f"Error: {e}"
