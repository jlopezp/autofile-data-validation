# ================================================================
# Archivo: validators/validacion_informe_Validacion_pagos.py
# Descripción:
#   Genera el informe final de pagos y el informe de validaciones.
#   Ahora muestra SIEMPRE las observaciones completas de cada validador.
#
# Autor: José López
# Institución: TRANSFORMACION DIGITAL
# Versión: v6.3 - Observaciones completas + soporte múltiple TXT
# Fecha: 17/11/2025
# ================================================================

import os
import pandas as pd
from datetime import datetime
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ================================================================
#  INFORME DE VALIDACIONES (CORREGIDO)
# ================================================================
def generar_informe_validaciones_pagos(resultados_validaciones, archivo_txt):
    """
    Genera un XLSX con:
      - Hoja 'Resumen Validaciones' con columnas (Validación, Resultado, Observaciones)
      - Hojas adicionales con detalle (si hay DataFrame de errores)
    Parámetros:
      - resultados_validaciones: lista de tuplas (nombre, resultado_texto, df_detalle_or_None)
      - archivo_txt: ruta del txt usado para ubicar carpeta 'Validaciones'
    """
    try:
        carpeta_base = os.path.dirname(archivo_txt)
        carpeta_validaciones = os.path.join(carpeta_base, "Validaciones")
        os.makedirs(carpeta_validaciones, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        nombre_salida = os.path.join(carpeta_validaciones, f"Validaciones_Pagos_{timestamp}.xlsx")

        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Validaciones"
        ws_resumen.append(["Validación", "Resultado", "Observaciones"])

        # Llevar control de nombres de hoja usados para evitar colisiones
        nombres_hoja_usados = set()

        for nombre, resultado_texto, df_detalle in resultados_validaciones:
            # Normalizar tipos
            if resultado_texto is None:
                resultado_texto = "OK"

            obs = str(resultado_texto)

            # Si el tercer elemento es DataFrame no vacío → agregar hoja con detalle
            if df_detalle is not None:
                try:
                    # Si viene como lista convertir a DataFrame
                    if isinstance(df_detalle, list):
                        df_detalle = pd.DataFrame(df_detalle, columns=["Observacion"])

                    if hasattr(df_detalle, "empty") and not df_detalle.empty:
                        filas = len(df_detalle)
                        obs = obs + f" — {filas} filas afectadas"

                        # construir nombre de hoja seguro (<=31 chars)
                        hoja_base = (nombre[:28] if nombre else "Detalle") .replace("/", "-")
                        hoja = hoja_base
                        suf = 1
                        while hoja in nombres_hoja_usados:
                            hoja = f"{hoja_base[:25]}_{suf}"
                            suf += 1
                        nombres_hoja_usados.add(hoja)

                        ws_det = wb.create_sheet(title=hoja)

                        # Escribir DataFrame en la hoja detalle (con encabezado)
                        for r in dataframe_to_rows(df_detalle, index=False, header=True):
                            ws_det.append(r)

                        # Aplicar una fila cabecera simple (opcional mejorable por estilo)
                        # (no se aplican estilos aquí para mantener simpleza)

                except Exception as e:
                    # Si falla convertir o escribir detalle, incluir la excepción en observaciones
                    obs = obs + f" — (No se pudo agregar detalle: {e})"

            # Añadir fila al resumen
            ws_resumen.append([nombre, resultado_texto, obs])

        # Guardar workbook
        wb.save(nombre_salida)
        print(f"✅ Informe de validaciones generado en: {nombre_salida}")

    except Exception as e:
        raise Exception(f"Error al generar el informe de validaciones: {e}")
