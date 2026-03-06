# ================================================================
# INFORME FINAL PAGOS – v7.3 ESTABLE (Formato TRANSFORMACION DIGITAL)
# ================================================================
def generar_informe_final_pagos(carpeta):
    import os
    import pandas as pd
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from tkinter import messagebox

    try:
        # ------------------------------------------------------------
        # 1) Validar carpeta
        # ------------------------------------------------------------
        if not carpeta or not os.path.isdir(carpeta):
            raise Exception("La carpeta recibida no es válida.")

        archivos = [f for f in os.listdir(carpeta) if f.lower().endswith(".txt")]
        if not archivos:
            raise Exception("No se encontraron archivos TXT en la carpeta seleccionada.")

        # ------------------------------------------------------------
        # 2) Leer TXT
        # ------------------------------------------------------------
        registros = []

        for archivo in archivos:
            ruta = os.path.join(carpeta, archivo)

            df = pd.read_csv(
                ruta,
                sep=";",
                header=None,
                dtype=str,
                encoding="latin-1",
                na_filter=False
            )

            if df.shape[1] != 12:
                print(f"Archivo ignorado (no tiene 12 columnas): {archivo}")
                continue

            df.columns = [
                "Rut", "Dv", "CodMoneda",
                "Operacion", "Cuotas", "Mes", "Año",
                "Dividendo", "Subsidio", "Subvencion",
                "Compensacion", "TotalPagar"
            ]

            # Entidad financiera desde nombre archivo
            nombre = os.path.splitext(archivo)[0]
            partes = nombre.split("_")

            entidad = "SinNombre"
            for p in partes:
                if p.lower() not in ["ds01", "ds02", "ds12", "ds51"]:
                    entidad = p
                    break

            df["EntidadFin"] = entidad

            # Convertir montos
            for col in ["Dividendo", "Subvencion", "Compensacion", "TotalPagar"]:
                df[col] = df[col].str.replace(",", ".", regex=False).astype(float)

            registros.append(df)

        if not registros:
            raise Exception("No existen datos válidos para consolidar.")

        # ------------------------------------------------------------
        # 3) Consolidar
        # ------------------------------------------------------------
        df_final = pd.concat(registros, ignore_index=True)
        df_final.insert(0, "NReg", df_final.index + 1)

        df_final = df_final[
            [
                "NReg",
                "EntidadFin",
                "Rut",
                "Dv",
                "CodMoneda",
                "Operacion",
                "Cuotas",
                "Mes",
                "Año",
                "Dividendo",
                "Subvencion",
                "Compensacion",
                "TotalPagar"
            ]
        ]

        # ------------------------------------------------------------
        # 4) Crear Excel
        # ------------------------------------------------------------
        salida = os.path.join(
            carpeta,
            f"Informe_Final_Pagos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe Final Pagos"

        headers = [
            "N° Reg",
            "Entidad Financiera",
            "Rut",
            "Dv",
            "Código Moneda Original",
            "N° de Operación",
            "N° Cuotas",
            "Mes",
            "Año",
            "Monto Dividendo en Moneda Original",
            "Monto Subvención en Moneda Original",
            "Compensación",
            "Total a Pagar"
        ]

        ws.append(headers)

        # Estilos
        header_fill = PatternFill(start_color="00142D", end_color="00142D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for _, row in df_final.iterrows():
            ws.append(list(row.values))

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.000000"

        # Total final
        fila_total = ws.max_row + 2
        ws.cell(row=fila_total, column=12, value="Total a Pagar").font = Font(bold=True)
        ws.cell(row=fila_total, column=13, value=df_final["TotalPagar"].sum()).font = Font(bold=True)
        ws.cell(row=fila_total, column=13).number_format = "#,##0.000000"

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) for c in col) + 2

        wb.save(salida)

        messagebox.showinfo("Informe Final Pagos", f"Informe generado correctamente:\n\n{salida}")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el informe final de pagos:\n{e}")
